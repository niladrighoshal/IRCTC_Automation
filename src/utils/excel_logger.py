import openpyxl
from openpyxl.styles import Font
import threading
from datetime import datetime
import os

class ExcelLogger:
    """
    A thread-safe logger for writing bot status updates to an Excel file using openpyxl.
    This is suitable for server environments without a display server.
    """
    def __init__(self, filename="irctc_booking_log.xlsx"):
        self.filename = filename
        self.lock = threading.Lock()
        # Initialize the file if it doesn't exist
        with self.lock:
            if not os.path.exists(self.filename):
                workbook = openpyxl.Workbook()
                # Rename the default sheet to 'Logs'
                sheet = workbook.active
                sheet.title = 'Logs'
                workbook.save(self.filename)

    def _get_column_number(self, instance_id):
        """Converts a 1-based instance ID to an Excel column number (1->2, 2->3, etc.)."""
        # Column 1 (A) is reserved for timestamps.
        return instance_id + 1

    def setup_column(self, instance_id, bot_config, config_filename):
        """
        Writes the static header information for a bot instance in its dedicated column.
        """
        with self.lock:
            try:
                workbook = openpyxl.load_workbook(self.filename)
                sheet = workbook['Logs']
                col = self._get_column_number(instance_id)
                row = 1

                account = bot_config.get('account', {})
                train_details = bot_config.get('train', {})
                passengers = bot_config.get('passengers', [])
                bold_font = Font(bold=True)

                data_to_log = {
                    "Browser Instance": instance_id,
                    "Using Config": os.path.basename(config_filename),
                    "Username": account.get('username', 'N/A'),
                    "Password": "*" * len(account.get('password', '')),
                    "Train No": train_details.get('train_no', 'N/A'),
                    "From": train_details.get('from_code', 'N/A'),
                    "To": train_details.get('to_code', 'N/A'),
                    "Date": datetime.strptime(train_details.get('date'), '%d%m%Y').strftime('%d-%b-%Y') if train_details.get('date') else 'N/A',
                    "Class": train_details.get('class', 'N/A'),
                    "Quota": train_details.get('quota', 'N/A'),
                }

                for key, value in data_to_log.items():
                    cell = sheet.cell(row=row, column=col, value=f"{key}: {value}")
                    cell.font = bold_font
                    row += 1

                cell = sheet.cell(row=row, column=col, value="--- Passengers ---")
                cell.font = bold_font
                row += 1
                for i, p in enumerate(passengers):
                    sheet.cell(row=row, column=col, value=f"P{i+1}: {p.get('name')}, {p.get('age')}, {p.get('sex')}")
                    row += 1

                row += 1
                cell_status = sheet.cell(row=row, column=col, value="--- Status Log ---")
                cell_status.font = bold_font
                cell_ts = sheet.cell(row=row, column=1, value="Timestamp")
                cell_ts.font = bold_font

                workbook.save(self.filename)
            except Exception as e:
                print(f"[ExcelLogger] Error setting up column for instance {instance_id}: {e}")

    def log(self, instance_id, message):
        """
        Logs a new status message for a given bot instance in the next available row.
        """
        with self.lock:
            try:
                workbook = openpyxl.load_workbook(self.filename)
                sheet = workbook['Logs']
                col = self._get_column_number(instance_id)

                # Find the last used row in that column to append the new log
                # openpyxl max_row is not always reliable for specific columns, so we iterate.
                new_row = 1
                while sheet.cell(row=new_row, column=col).value is not None:
                    new_row += 1

                sheet.cell(row=new_row, column=1, value=datetime.now().strftime('%H:%M:%S.%f')[:-3])
                sheet.cell(row=new_row, column=col, value=message)

                workbook.save(self.filename)
            except Exception as e:
                print(f"[ExcelLogger] Error logging for instance {instance_id}: {e}")
